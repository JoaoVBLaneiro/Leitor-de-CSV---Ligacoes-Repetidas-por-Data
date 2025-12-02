import streamlit as st
import pandas as pd
from fpdf import FPDF
import io

st.set_page_config(page_title="Analisador de Telefones", layout="centered")

st.title("📊 Analisador de Telefones por Data")
st.write("Envie arquivos CSV contendo as colunas **Data/Hora** e **Origem**.")

# Upload de múltiplos arquivos
uploaded_files = st.file_uploader(
    "Selecione os arquivos CSV",
    type=["csv"],
    accept_multiple_files=True
)

if uploaded_files:
    lista_df = []

    for file in uploaded_files:
        df = pd.read_csv(file)
        df.columns = df.columns.str.lower()  # normaliza nomes

        # Identifica colunas Data/Hora e Origem
        col_data = None
        col_origem = None
        for c in df.columns:
            if "data" in c:
                col_data = c
            if "origem" in c:
                col_origem = c

        if col_data is None or col_origem is None:
            st.error(f"🚫 Arquivo ignorado: {file.name} (faltam Data/Hora ou Origem)")
            continue

        # Converte Data/Hora para data
        df[col_data] = pd.to_datetime(df[col_data], errors="coerce")
        df["data"] = df[col_data].dt.date

        # Prepara df com colunas padronizadas
        df = df[["data", col_origem]].rename(columns={col_origem: "telefone"})
        lista_df.append(df)

    if lista_df:
        # Concatena todos os CSVs
        dados = pd.concat(lista_df, ignore_index=True)

        # Agrupamento e contagem
        agrupado = dados.groupby(["data", "telefone"]).size().reset_index(name="ocorrencias")

        # Pivot para cada data virar coluna
        tabela_final = agrupado.pivot_table(
            index="telefone",
            columns="data",
            values="ocorrencias",
            fill_value=0
        )

        # Coluna Total
        tabela_final["Total"] = tabela_final.sum(axis=1)
        tabela_final = tabela_final.sort_values("Total", ascending=False)

        st.subheader("Resultado Consolidado (por número e por data)")
        st.dataframe(tabela_final, use_container_width=True)

        # Exportar CSV
        csv = tabela_final.to_csv().encode("utf-8")
        st.download_button(
            "📥 Baixar Resultado (CSV)",
            csv,
            "resultado_pivotado.csv",
            "text/csv"
        )

        # Exportar XLSX
        try:
            import openpyxl
            output = io.BytesIO()
            tabela_final.to_excel(output, index=True, engine='openpyxl')
            st.download_button(
                "📥 Baixar Resultado (XLSX)",
                data=output.getvalue(),
                file_name="resultado_pivotado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except ImportError:
            st.warning("Para exportar em XLSX, instale a biblioteca openpyxl: pip install openpyxl")

        # Exportar PDF
        def gerar_pdf(df):
            
            # Converte números para string sem decimais
            df_pdf = df.copy()
            for col in df_pdf.columns:
                if pd.api.types.is_numeric_dtype(df_pdf[col]):
                    df_pdf[col] = df_pdf[col].astype(int).astype(str)

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Relatório de Ligações", ln=True, align="C")
            pdf.ln(10)

            pdf.set_font("Arial", "", 10)

            # Cabeçalho
            colunas = df.columns.tolist()
            for col in colunas:
                pdf.cell(30, 8, str(col), border=1, align="C")
            pdf.ln()

            # Linhas do DataFrame
            for _, row in df.iterrows():
                for col in colunas:
                    pdf.cell(30, 8, str(row[col]), border=1, align="C")
                pdf.ln()

            # Gera PDF como bytes
            pdf_bytes = pdf.output(dest='S').encode('latin1')  # 'S' retorna string em bytes
            return io.BytesIO(pdf_bytes)

        
        pdf_bytes = gerar_pdf(tabela_final.reset_index())
        st.download_button(
            "📥 Baixar Resultado (PDF)",
            data=pdf_bytes,
            file_name="resultado_pivotado.pdf",
            mime="application/pdf"
        )


    else:
        st.warning("Nenhum arquivo válido foi enviado.")
